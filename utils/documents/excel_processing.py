import openpyxl

from config import FLOAT_NUMBER_REG
from storage.storage_handler import StorageHandler

storage_handler = StorageHandler()

import openpyxl
import re


def process_excel_on_upload(file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Init returned values
    total_sum = 0.0
    excel_text = ""

    # Loop through all sheets in the workbook
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Loop through all rows and columns to extract text
        for row in worksheet.iter_rows():
            for cell in row:
                cell_value = cell.value
                if isinstance(cell_value, str):
                    # Extract all numeric values (including decimals) using regular expressions
                    numbers = re.findall(FLOAT_NUMBER_REG, cell_value)
                    for num in numbers:
                        total_sum += float(num)
                elif isinstance(cell_value, float) or isinstance(cell_value, int):
                    total_sum += cell_value

                excel_text += f"{cell_value} "


    return excel_text, total_sum